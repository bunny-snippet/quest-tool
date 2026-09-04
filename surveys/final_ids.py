"""Final-ID reconciliation import, validation and audit persistence."""

from __future__ import annotations

import csv
import hashlib
import io
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from vendors.models import Client

from .models import FinalIDStatus, FinalIDUpload, FinalIDUploadItem, SurveyAttempt


MAX_UPLOAD_BYTES = 15 * 1024 * 1024
MAX_UNIQUE_RIDS = 50_000
RID_HEADERS = {"rid", "respondentid", "respondentidentifier"}


class FinalIDImportError(ValueError):
    """A safe, user-facing validation failure for a final-ID upload."""


def _normalise_header(value) -> str:
    return "".join(character for character in str(value or "").lower() if character.isalnum())


def _normalise_rid(value) -> str:
    rid = str(value or "").strip()
    # Excel users commonly prefix identifiers with an apostrophe to force text
    # formatting. The apostrophe is not part of the platform RID.
    if rid.startswith("'") and len(rid) == 11:
        rid = rid[1:]
    return rid


def _csv_values(uploaded_file) -> Iterable[object]:
    text_stream = None
    try:
        uploaded_file.seek(0)
        text_stream = io.TextIOWrapper(uploaded_file.file, encoding="utf-8-sig", newline="")
        rows = csv.reader(text_stream)
        headers = next(rows, None)
    except (UnicodeDecodeError, csv.Error) as exc:
        raise FinalIDImportError("CSV must be a valid UTF-8 file with an RID column.") from exc
    if not headers:
        raise FinalIDImportError("The CSV file is empty.")
    rid_index = next(
        (index for index, value in enumerate(headers) if _normalise_header(value) in RID_HEADERS),
        None,
    )
    if rid_index is None:
        raise FinalIDImportError("The file must contain an RID column.")
    try:
        for row in rows:
            if rid_index < len(row):
                yield row[rid_index]
    finally:
        if text_stream is not None:
            text_stream.detach()


def _xlsx_values(uploaded_file) -> Iterable[object]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise FinalIDImportError("Excel import is temporarily unavailable. Please upload CSV instead.") from exc

    try:
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
    except Exception as exc:
        raise FinalIDImportError("Excel file could not be read. Upload a valid .xlsx file.") from exc
    if not headers:
        workbook.close()
        raise FinalIDImportError("The Excel file is empty.")
    rid_index = next(
        (index for index, value in enumerate(headers) if _normalise_header(value) in RID_HEADERS),
        None,
    )
    if rid_index is None:
        workbook.close()
        raise FinalIDImportError("The file must contain an RID column.")
    try:
        for row in rows:
            if rid_index < len(row):
                yield row[rid_index]
    finally:
        workbook.close()


def read_uploaded_rids(uploaded_file) -> tuple[list[str], int, int, str]:
    """Return distinct valid RIDs, row counts and the immutable file digest."""

    filename = Path(str(uploaded_file.name or "")).name
    suffix = Path(filename).suffix.lower()
    size = int(getattr(uploaded_file, "size", 0) or 0)
    if not filename or suffix not in {".csv", ".xlsx"}:
        raise FinalIDImportError("Upload a CSV or Excel (.xlsx) file.")
    if size < 1:
        raise FinalIDImportError("The selected file is empty.")
    if size > MAX_UPLOAD_BYTES:
        raise FinalIDImportError("The selected file is larger than the 15 MB limit.")

    hasher = hashlib.sha256()
    for chunk in uploaded_file.chunks():
        hasher.update(chunk)
    values = _csv_values(uploaded_file) if suffix == ".csv" else _xlsx_values(uploaded_file)

    submitted_count = 0
    rids: list[str] = []
    seen: set[str] = set()
    invalid_rows = 0
    for value in values:
        rid = _normalise_rid(value)
        if not rid:
            continue
        submitted_count += 1
        if len(rid) != 10 or not rid.isalnum():
            invalid_rows += 1
            continue
        if rid not in seen:
            seen.add(rid)
            rids.append(rid)
        if len(rids) > MAX_UNIQUE_RIDS:
            raise FinalIDImportError("One upload can contain at most 50,000 unique RIDs.")
    if not rids:
        if invalid_rows:
            raise FinalIDImportError(
                f"No valid 10-character RIDs were found; {invalid_rows} invalid row(s) were skipped."
            )
        raise FinalIDImportError("No RID values were found in the uploaded file.")
    return rids, submitted_count, invalid_rows, hasher.hexdigest()


def _attempt_matches_client(attempt: SurveyAttempt, client_id: int) -> bool:
    survey = attempt.survey
    return client_id in {
        attempt.client_id,
        survey.client_id,
        survey.integration.client_id if survey.integration_id else None,
    }


def _client_attempt_filter(client_id: int) -> Q:
    """Match the client snapshot, survey client or integration client."""

    return (
        Q(client_id=client_id)
        | Q(survey__client_id=client_id)
        | Q(survey__integration__client_id=client_id)
    )


def _accounting_month_bounds(accounting_month: date) -> tuple[datetime, datetime]:
    if accounting_month.month == 12:
        next_month = date(accounting_month.year + 1, 1, 1)
    else:
        next_month = date(accounting_month.year, accounting_month.month + 1, 1)
    timezone_value = timezone.get_current_timezone()
    return (
        timezone.make_aware(datetime.combine(accounting_month, datetime.min.time()), timezone_value),
        timezone.make_aware(datetime.combine(next_month, datetime.min.time()), timezone_value),
    )


def import_final_ids(
    *,
    uploaded_file,
    client: Client,
    accounting_month: date,
    decision: str,
    uploaded_by,
) -> dict:
    """Apply one client final-ID file without changing survey lifecycle status.

    The selected accounting month belongs to the upload itself. RID lookup is
    intentionally global, then constrained to the selected client, so a later
    reconciliation file can update an older journey while its revenue remains
    attributable to the selected invoice month.
    """

    if decision not in FinalIDUpload.Decision.values:
        raise FinalIDImportError("Choose Accepted or Rejected.")
    if accounting_month.day != 1:
        raise FinalIDImportError("Accounting month must use the first day of the month.")

    rids, submitted_count, invalid_count, file_sha256 = read_uploaded_rids(uploaded_file)
    filename = Path(str(uploaded_file.name or "")).name
    now = timezone.now()

    with transaction.atomic():
        upload = FinalIDUpload.objects.create(
            client=client,
            accounting_month=accounting_month,
            decision=decision,
            original_filename=filename,
            file_sha256=file_sha256,
            uploaded_by=uploaded_by,
            submitted_count=submitted_count,
            unique_rid_count=len(rids),
            invalid_count=invalid_count,
        )
        attempts = {
            attempt.rid: attempt
            for attempt in SurveyAttempt.objects.select_for_update().select_related(
                "survey__integration", "client"
            ).filter(rid__in=rids)
        }
        final_lifecycle_statuses = {
            SurveyAttempt.Status.COMPLETED,
            SurveyAttempt.Status.TERMINATED,
            SurveyAttempt.Status.OVER_QUOTA,
            SurveyAttempt.Status.QUALITY_TERMINATED,
        }
        eligible_attempts = {
            rid: attempt
            for rid, attempt in attempts.items()
            if _attempt_matches_client(attempt, client.pk)
            and attempt.status in final_lifecycle_statuses
        }

        # An Accepted file is the client's authoritative completed-RID list for
        # the selected activity month. Completed journeys omitted from that
        # file become Client Rejected, while directly supplied RIDs can belong
        # to any activity month and always retain the selected invoice month.
        auto_rejected_attempts = []
        if decision == FinalIDUpload.Decision.ACCEPTED:
            lower, upper = _accounting_month_bounds(accounting_month)
            auto_rejected_attempts = list(
                SurveyAttempt.objects.select_for_update().select_related(
                    "survey__integration", "client"
                ).filter(
                    _client_attempt_filter(client.pk),
                    status=SurveyAttempt.Status.COMPLETED,
                    initiated_at__gte=lower,
                    initiated_at__lt=upper,
                ).exclude(rid__in=rids)
            )
        all_eligible_attempts = [
            *eligible_attempts.values(),
            *auto_rejected_attempts,
        ]
        existing_statuses = {
            item.attempt_id: item
            for item in FinalIDStatus.objects.select_for_update().filter(
                attempt_id__in=[attempt.pk for attempt in all_eligible_attempts]
            )
        }

        status_creates = []
        status_updates = []
        upload_items = []
        counters = {
            "applied": 0,
            "not_found": 0,
            "client_mismatch": 0,
            "not_completed": 0,
            "auto_rejected": 0,
        }

        def apply_status(attempt, applied_status, rid):
            previous = existing_statuses.get(attempt.pk)
            previous_status = previous.status if previous else ""
            if previous is None:
                status = FinalIDStatus(
                    attempt=attempt,
                    client=client,
                    status=applied_status,
                    accounting_month=accounting_month,
                    upload=upload,
                )
                status_creates.append(status)
                existing_statuses[attempt.pk] = status
            else:
                previous.client = client
                previous.status = applied_status
                previous.accounting_month = accounting_month
                previous.upload = upload
                previous.updated_at = now
                status_updates.append(previous)
            upload_items.append(FinalIDUploadItem(
                upload=upload,
                rid=rid,
                attempt=attempt,
                outcome=FinalIDUploadItem.Outcome.APPLIED,
                previous_status=previous_status,
                applied_status=applied_status,
            ))

        for rid in rids:
            attempt = attempts.get(rid)
            if attempt is None:
                outcome = FinalIDUploadItem.Outcome.NOT_FOUND
                counters["not_found"] += 1
                upload_items.append(FinalIDUploadItem(upload=upload, rid=rid, outcome=outcome))
                continue
            if not _attempt_matches_client(attempt, client.pk):
                counters["client_mismatch"] += 1
                upload_items.append(FinalIDUploadItem(
                    upload=upload, rid=rid, attempt=attempt,
                    outcome=FinalIDUploadItem.Outcome.CLIENT_MISMATCH,
                ))
                continue
            if attempt.status not in final_lifecycle_statuses:
                counters["not_completed"] += 1
                upload_items.append(FinalIDUploadItem(
                    upload=upload, rid=rid, attempt=attempt,
                    outcome=FinalIDUploadItem.Outcome.NOT_COMPLETED,
                ))
                continue

            counters["applied"] += 1
            apply_status(attempt, decision, rid)

        for attempt in auto_rejected_attempts:
            counters["auto_rejected"] += 1
            apply_status(attempt, FinalIDUpload.Decision.REJECTED, attempt.rid)

        FinalIDStatus.objects.bulk_create(status_creates, batch_size=1000)
        if status_updates:
            FinalIDStatus.objects.bulk_update(
                status_updates,
                ["client", "status", "accounting_month", "upload", "updated_at"],
                batch_size=1000,
            )
        FinalIDUploadItem.objects.bulk_create(upload_items, batch_size=1000)
        FinalIDUpload.objects.filter(pk=upload.pk).update(
            applied_count=counters["applied"],
            not_found_count=counters["not_found"],
            client_mismatch_count=counters["client_mismatch"],
            not_completed_count=counters["not_completed"],
            auto_rejected_count=counters["auto_rejected"],
        )

    return {
        "upload_id": upload.pk,
        "submitted": submitted_count,
        "unique": len(rids),
        "invalid": invalid_count,
        **counters,
    }
